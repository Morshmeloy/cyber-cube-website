import io
from datetime import datetime
from typing import Any
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

ORGANIZATION_NAME = 'ООО "Д4 ТЕХНОЛОГИИ"'
WAREHOUSE_NAME = "Основной склад"


def _write_rows(headers: list[str], rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_operations_export(operations: list) -> bytes:
    """Экспорт истории выдач/возвратов — для переноса записей в 1С вручную."""
    rows = [
        [
            op.created_at.strftime("%d.%m.%Y %H:%M"),
            op.nomenclature.name,
            "Выдача" if op.operation_type.value == "issue" else "Возврат",
            op.quantity,
            op.person,
            op.destination,
            op.user.username,
        ]
        for op in operations
    ]
    return _write_rows(
        [
            "Дата",
            "Номенклатура",
            "Тип",
            "Количество",
            "ФИО",
            "Адрес/место назначения",
            "Кто ввёл",
        ],
        rows,
    )


_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")

_FONT_TITLE = Font(name="Arial", size=14, bold=True)
_FONT_LABEL = Font(name="Arial", size=10, bold=False)
_FONT_VALUE = Font(name="Arial", size=10, bold=True)
_FONT_HEADER = Font(name="Arial", size=9, bold=True)
_FONT_DATA = Font(name="Arial", size=8, bold=False)
_FONT_SIGN_LABEL = Font(name="Arial", size=9, bold=True)
_FONT_SIGN_VALUE = Font(name="Arial", size=10, bold=False)

_ALIGN_TITLE = Alignment(horizontal="left", vertical="center")
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center")
_ALIGN_NUM = Alignment(horizontal="right", vertical="top")
_ALIGN_TEXT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_ALIGN_LABEL = Alignment(horizontal="left", vertical="center")
_ALIGN_SIGN_LABEL = Alignment(horizontal="right")

# Раскладка колонок повторяет присланный образец: A — узкий отступ слева,
# данные — в B..K (№ / Код(C:D) / Материал(E:H) / Количество(I) / Ед.изм.(J:K,
# без отдельного заголовка — обе подписаны одним общим "Количество" в шапке).
_COLUMN_WIDTHS = {
    "A": 1,
    "B": 6.33,
    "C": 4.66,
    "D": 12,
    "E": 19.5,
    "F": 10.66,
    "G": 19.83,
    "H": 10,
    "I": 13.16,
    "J": 27,
    "K": 0.16,
}


def _set_cell(sheet, coord: str, value=None, font=None, alignment=None):
    cell = sheet[coord]
    if value is not None:
        cell.value = value
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    return cell


def build_requirement_invoice_export(
    operations: list,
    *,
    invoice_number: str | None = None,
    contract_name: str | None = None,
    released_by: str | None = None,
    received_by: str | None = None,
) -> bytes:
    """Экспорт выбранных операций в формате «Требование-накладная» — раскладка,
    шрифты и границы сняты 1-в-1 с присланного бланка-образца (см. корень репозитория)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Требование-накладная"

    for col, width in _COLUMN_WIDTHS.items():
        sheet.column_dimensions[col].width = width

    today = datetime.now().strftime("%d.%m.%Y")

    # 1. Заголовок — подчёркнут снизу (medium), не по центру, а слева
    sheet.merge_cells("B3:K3")
    number = invoice_number or "_____"
    _set_cell(sheet, "B3", f"Требование-накладная № {number} от {today} г.", _FONT_TITLE, _ALIGN_TITLE)
    for col in "BCDEFGHIJK":
        sheet[f"{col}3"].border = Border(bottom=_MEDIUM)

    # 2. Организация / Склад — подпись обычным, значение жирным (как в образце)
    _set_cell(sheet, "B5", "Организация:", _FONT_LABEL, _ALIGN_LABEL)
    sheet.merge_cells("E5:K5")
    _set_cell(sheet, "E5", ORGANIZATION_NAME, _FONT_VALUE, Alignment(horizontal="left", wrap_text=True))

    _set_cell(sheet, "B8", "Склад:", _FONT_LABEL, _ALIGN_LABEL)
    sheet.merge_cells("E8:K8")
    _set_cell(sheet, "E8", WAREHOUSE_NAME, _FONT_VALUE, Alignment(horizontal="left", wrap_text=True))

    # 3. Договор с заказчиком — вписывается вручную при формировании документа.
    # Объект — своё значение уже есть на каждой операции (destination), в бланке-
    # образце для него не было отдельной ячейки, добавляем её тем же стилем, что у
    # Организации/Склада. Подпись "Договор с заказчиком" длиннее остальных подписей,
    # поэтому колонка D расширена (см. _COLUMN_WIDTHS) — значение остаётся на одном
    # уровне с Организацией/Складом/Объектом (с колонки E), но подпись уже не наезжает.
    _set_cell(sheet, "B9", "Договор с заказчиком", _FONT_LABEL, _ALIGN_LABEL)
    if contract_name:
        sheet.merge_cells("E9:K9")
        _set_cell(sheet, "E9", contract_name, _FONT_VALUE, Alignment(horizontal="left", wrap_text=True))

    _set_cell(sheet, "B10", "Объект", _FONT_LABEL, _ALIGN_LABEL)
    object_name = ", ".join(sorted({op.destination for op in operations if getattr(op, "destination", None)}))
    if object_name:
        sheet.merge_cells("E10:K10")
        _set_cell(sheet, "E10", object_name, _FONT_VALUE, Alignment(horizontal="left", wrap_text=True))

    # 4. Заголовки таблицы: № / Код(C:D) / Материал(E:H) / Количество(I:K,
    # без отдельного заголовка на единицу измерения)
    sheet.merge_cells("C12:D12")
    sheet.merge_cells("E12:H12")
    sheet.merge_cells("I12:K12")
    _set_cell(sheet, "B12", "№", _FONT_HEADER, _ALIGN_HEADER)
    _set_cell(sheet, "C12", "Код", _FONT_HEADER, _ALIGN_HEADER)
    _set_cell(sheet, "E12", "Материал", _FONT_HEADER, _ALIGN_HEADER)
    _set_cell(sheet, "I12", "Количество", _FONT_HEADER, _ALIGN_HEADER)
    for col in "BCDEFGHIJK":
        cell = sheet[f"{col}12"]
        cell.border = Border(
            top=_MEDIUM,
            bottom=_THIN,
            left=_MEDIUM if col == "B" else (_THIN if col == "I" else None),
            right=_MEDIUM if col == "K" else (_THIN if col in "BDH" else None),
        )

    # 5. Данные — по строке на позицию
    row = 13
    for i, op in enumerate(operations, start=1):
        if not getattr(op, "nomenclature", None):
            continue

        sheet.merge_cells(f"C{row}:D{row}")
        sheet.merge_cells(f"E{row}:H{row}")
        sheet.merge_cells(f"J{row}:K{row}")

        _set_cell(sheet, f"B{row}", i, _FONT_DATA, _ALIGN_NUM)
        _set_cell(sheet, f"C{row}", op.nomenclature.code or "", _FONT_DATA, _ALIGN_TEXT_WRAP)
        _set_cell(sheet, f"E{row}", op.nomenclature.name, _FONT_DATA, _ALIGN_TEXT_WRAP)
        _set_cell(sheet, f"I{row}", op.quantity, _FONT_DATA, _ALIGN_NUM)
        _set_cell(sheet, f"J{row}", op.nomenclature.unit or "", _FONT_DATA, _ALIGN_NUM)

        for col in "BCDEFGHIJK":
            cell = sheet[f"{col}{row}"]
            cell.border = Border(
                top=_THIN,
                bottom=_THIN,
                left=_MEDIUM if col == "B" else _THIN,
                right=_MEDIUM if col == "K" else _THIN,
            )
        row += 1

    # 6. Нижняя граница таблицы (медиум, во всю ширину)
    for col in "BCDEFGHIJK":
        sheet[f"{col}{row}"].border = Border(top=_MEDIUM)
    row += 2

    # 7. Подписи "Отпустил" / "Получил" — подпись справа от метки, ФИО (если указано
    # при формировании документа) печатается на самой линии, место для росписи от
    # руки остаётся тем же — в образце ФИО не печаталось, линия была пустой.
    sheet.merge_cells(f"B{row}:C{row}")
    _set_cell(sheet, f"B{row}", "Отпустил", _FONT_SIGN_LABEL, _ALIGN_SIGN_LABEL)
    sheet.merge_cells(f"D{row}:F{row}")
    if released_by:
        _set_cell(sheet, f"D{row}", released_by, _FONT_SIGN_VALUE, Alignment(horizontal="center", vertical="bottom"))
    for col in "DEF":
        sheet[f"{col}{row}"].border = Border(bottom=_THIN)

    _set_cell(sheet, f"H{row}", "Получил", _FONT_SIGN_LABEL, _ALIGN_SIGN_LABEL)
    sheet.merge_cells(f"I{row}:J{row}")
    if received_by:
        _set_cell(sheet, f"I{row}", received_by, _FONT_SIGN_VALUE, Alignment(horizontal="center", vertical="bottom"))
    for col in "IJ":
        sheet[f"{col}{row}"].border = Border(bottom=_THIN)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
